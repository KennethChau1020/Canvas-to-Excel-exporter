import os
import re
import time
import colorsys
from datetime import datetime, date
from urllib.parse import urljoin

import requests
from ics import Calendar

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

# IMPORTANT:
# Change this to your school's Canvas address.
#
# Examples:
# https://canvas.school.edu
# https://myschool.instructure.com
#
CANVAS_BASE_URL = "https://canvas.instructure.com"


# ------------------------------------------------------------
# HEADLESS MODE
#
# False = Chrome is visible.
# True  = Chrome runs invisibly.
#
# I recommend using False while testing.
#
# If you later use an existing Chrome profile so Canvas remains
# signed in, you can change this to True.
# ------------------------------------------------------------
HEADLESS = False


CHROME_USER_DATA_DIR = None
CHROME_PROFILE_DIRECTORY = "Default"


# Pages checked by Selenium.
COURSE_TABS = [
    "assignments",
    "quizzes",
    "modules"
]


# Items we deliberately do not want in the spreadsheet.
SKIP_WORDS = [
    "syllabus",
    "roll call",
    "attendance"
]


# GENERAL HELPER FUNCTIONS


def normalize_text(text):
    """
    Normalize assignment/course text for duplicate checking.
    """
    if not text:
        return ""

    text = text.lower().strip()

    # Replace multiple spaces with one.
    text = re.sub(r"\s+", " ", text)

    # Remove punctuation that should not matter for duplicate matching.
    text = re.sub(r"[^\w\s-]", "", text)

    return text


def parse_user_date(prompt):
    """
    Repeatedly ask for a date in:
        April 07 2025
    format.
    """
    while True:
        value = input(prompt).strip()

        try:
            return datetime.strptime(value, "%B %d %Y").date()

        except ValueError:
            print(
                "Invalid date format.\n"
                "Please enter the date like:\n"
                "April 07 2025"
            )


def assignment_key(assignment):
    """
    Produce a comparison key used for removing duplicates.
    """
    return (
        normalize_text(assignment.get("Assignment", "")),
        normalize_text(assignment.get("Course", ""))
    )


def clean_filename(filename):
    """
    Make sure the output filename ends in .xlsx.
    """
    filename = filename.strip()

    if not filename:
        filename = "canvas_assignments"

    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    return filename

# DATE PARSING


MONTH_PATTERN = (
    r"(January|February|March|April|May|June|July|August|"
    r"September|October|November|December|"
    r"Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)"
)


def extract_date_from_text(text, assumed_year):
    """
    Try to find a Canvas-style date inside arbitrary text.

    Examples:
        Apr 8
        April 08
        Apr 8 at 11:59pm
        Due Apr 8
        Available until April 8
    """

    if not text:
        return None

    pattern = rf"{MONTH_PATTERN}\s+(\d{{1,2}})"

    match = re.search(pattern, text, re.IGNORECASE)

    if not match:
        return None

    month_text = match.group(1)
    day_text = match.group(2)

    # Add the year BEFORE calling strptime.
    # This avoids Python's missing-year / leap-year warning.
    full_date_text = f"{month_text} {day_text} {assumed_year}"

    formats = [
        "%B %d %Y",
        "%b %d %Y"
    ]

    # Python recognizes Sep but not always Sept.
    full_date_text = re.sub(
        r"\bSept\b",
        "Sep",
        full_date_text,
        flags=re.IGNORECASE
    )

    for fmt in formats:
        try:
            return datetime.strptime(full_date_text, fmt).date()
        except ValueError:
            pass

    return None

# ICS CALENDAR


def extract_ics_course_and_assignment(title):
    """
    Canvas ICS titles often look like:

        Homework 4 [25S_CE2051.01]
        Quiz 2 [BIO101]

    Pull the square-bracket course text into the Course column.
    """

    title = title.strip()

    match = re.search(r"\[([^\]]+)\]", title)

    if match:
        course = match.group(1).strip()

        assignment = (
            title[:match.start()] +
            title[match.end():]
        ).strip()

    else:
        course = "Unknown"
        assignment = title

    return assignment, course


def parse_ics_calendar(ics_source, start_date, end_date):
    """
    Download/read the ICS calendar and return assignments within
    the requested date range.
    """

    print("\nReading Canvas ICS calendar...")

    try:
        if ics_source.lower().startswith(("http://", "https://")):

            response = requests.get(
                ics_source,
                timeout=30
            )

            response.raise_for_status()

            calendar = Calendar(response.text)

        else:

            with open(
                ics_source,
                "r",
                encoding="utf-8"
            ) as file:

                calendar = Calendar(file.read())

    except Exception as error:
        print("\nThe ICS calendar could not be read.")
        print(error)
        return []

    assignments = []

    for event in calendar.events:

        try:
            if not event.begin:
                continue

            event_datetime = event.begin.datetime
            event_date = event_datetime.date()

            if not (start_date <= event_date <= end_date):
                continue

            assignment_name, course_name = (
                extract_ics_course_and_assignment(
                    event.name or ""
                )
            )

            # Skip things like Syllabus / Roll Call.
            lower_name = assignment_name.lower()

            if any(
                skip_word in lower_name
                for skip_word in SKIP_WORDS
            ):
                continue

            assignments.append(
                {
                    "Assignment": assignment_name,
                    "Course": course_name,
                    "Due Date": event_date,
                    "Notes": ""
                }
            )

        except Exception:
            continue

    print(
        f"ICS scan found {len(assignments)} "
        f"calendar items in the selected date range."
    )

    return assignments


# SELENIUM SETUP


def create_driver():
    """
    Create the Chrome Selenium driver.

    Modern Selenium normally downloads/manages ChromeDriver
    automatically, so a manually downloaded chromedriver.exe
    usually isn't necessary anymore.
    """

    options = webdriver.ChromeOptions()

    if HEADLESS:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
    else:
        options.add_argument("--start-maximized")

    if CHROME_USER_DATA_DIR:
        options.add_argument(
            f"--user-data-dir={CHROME_USER_DATA_DIR}"
        )

        options.add_argument(
            f"--profile-directory={CHROME_PROFILE_DIRECTORY}"
        )

    options.add_argument("--disable-notifications")

    driver = webdriver.Chrome(options=options)

    return driver



# CANVAS COURSE LIST


def get_canvas_courses(driver):
    """
    Attempt to retrieve the student's Canvas course list.

    First tries Dashboard Cards, then falls back to /courses.
    """

    print("\nReading your Canvas courses...")

    courses = []

    # Try Dashboard cards first.

    driver.get(
        CANVAS_BASE_URL.rstrip("/") + "/"
    )

    try:

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.TAG_NAME, "body")
            )
        )

    except Exception:
        pass

    cards = driver.find_elements(
        By.CSS_SELECTOR,
        ".ic-DashboardCard"
    )

    for card in cards:

        try:

            link = card.find_element(
                By.CSS_SELECTOR,
                "a[href*='/courses/']"
            )

            href = link.get_attribute("href")

            # Try several possible title locations.
            name = ""

            selectors = [
                ".ic-DashboardCard__header-title",
                ".ic-DashboardCard__header-subtitle",
                "span"
            ]

            for selector in selectors:

                try:

                    element = card.find_element(
                        By.CSS_SELECTOR,
                        selector
                    )

                    possible_name = element.text.strip()

                    if possible_name:
                        name = possible_name
                        break

                except Exception:
                    pass

            if name and href:

                courses.append(
                    (name, href)
                )

        except Exception:
            continue

    # Fallback: All Courses page.


    if not courses:

        driver.get(
            CANVAS_BASE_URL.rstrip("/") + "/courses"
        )

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.TAG_NAME, "body")
            )
        )

        time.sleep(2)

        links = driver.find_elements(
            By.CSS_SELECTOR,
            "a[href*='/courses/']"
        )

        for link in links:

            try:

                name = link.text.strip()
                href = link.get_attribute("href")

                if not name or not href:
                    continue

                # We only want direct course links.
                course_match = re.search(
                    r"/courses/(\d+)/?$",
                    href
                )

                if not course_match:
                    continue

                courses.append(
                    (name, href)
                )

            except Exception:
                continue

    # Remove duplicate course links.
    unique_courses = []
    seen_urls = set()

    for course_name, course_url in courses:

        clean_url = course_url.rstrip("/")

        if clean_url in seen_urls:
            continue

        seen_urls.add(clean_url)

        unique_courses.append(
            (
                course_name,
                clean_url
            )
        )

    return unique_courses


# COURSE SELECTION MENU


def select_courses(courses):
    """
    Let the user toggle multiple courses.

    Example:
        1 3 4

    0 ends selection.
    """

    if not courses:
        print("\nNo Canvas courses could be found.")
        return []

    selected = set()

    number_width = len(str(len(courses)))

    while True:

        print("\nYour Courses:\n")

        longest_course = max(
            len(name)
            for name, _ in courses
        )

        for index, (course_name, _) in enumerate(
            courses,
            start=1
        ):

            status = (
                "SELECTED"
                if index in selected
                else "NOT SELECTED"
            )

            print(
                f"Class {index:>{number_width}}: "
                f"{course_name:<{longest_course + 5}}"
                f"{status}"
            )

        print(
            "\nPlease separate multiple course numbers "
            "using spaces."
        )

        value = input(
            "Type course numbers to toggle "
            "(or 0 when finished): "
        ).strip()

        entries = value.split()

        if "0" in entries:
            break

        for entry in entries:

            if not entry.isdigit():
                print(
                    f"Ignoring invalid entry: {entry}"
                )
                continue

            number = int(entry)

            if not (1 <= number <= len(courses)):
                print(
                    f"Class {number} does not exist."
                )
                continue

            if number in selected:
                selected.remove(number)
            else:
                selected.add(number)

    return [
        courses[index - 1]
        for index in sorted(selected)
    ]



def candidate_assignment_links(driver):
    """
    Dynamically collect links that look like Canvas graded work.

    This deliberately uses URL patterns instead of depending
    completely on fragile CSS classes such as ig-row.
    """

    selectors = [
        "a[href*='/assignments/']",
        "a[href*='/quizzes/']",
        "a[href*='/discussion_topics/']"
    ]

    found = {}

    for selector in selectors:

        links = driver.find_elements(
            By.CSS_SELECTOR,
            selector
        )

        for link in links:

            try:

                href = link.get_attribute("href")

                if not href:
                    continue

                # Exclude things such as submissions/grades.
                lower_href = href.lower()

                if any(
                    unwanted in lower_href
                    for unwanted in [
                        "/submissions/",
                        "/grades",
                        "/statistics"
                    ]
                ):
                    continue

                name = (
                    link.text
                    or link.get_attribute("aria-label")
                    or link.get_attribute("title")
                    or ""
                ).strip()

                if not name:
                    continue

                lower_name = name.lower()

                if any(
                    skip_word in lower_name
                    for skip_word in SKIP_WORDS
                ):
                    continue

                # Get surrounding visible text BEFORE navigating,
                # because Selenium elements become stale afterward.
                context_text = ""

                ancestor_paths = [
                    "./ancestor::li[1]",
                    "./ancestor::div[1]",
                    "./parent::*"
                ]

                for ancestor_path in ancestor_paths:

                    try:

                        ancestor = link.find_element(
                            By.XPATH,
                            ancestor_path
                        )

                        candidate_context = (
                            ancestor.text.strip()
                        )

                        if len(candidate_context) > len(context_text):
                            context_text = candidate_context

                    except Exception:
                        pass

                # URL is a convenient natural duplicate key.
                found[href] = {
                    "name": name,
                    "href": href,
                    "context": context_text
                }

            except Exception:
                continue

    return list(found.values())


# POINTS CHECK


def has_points(text):
    """
    Look for common Canvas point displays.

    Examples:
        20 pts
        20 points
        Points 20
        100 pts possible
    """

    if not text:
        return False

    patterns = [
        r"\b\d+(?:\.\d+)?\s*pts?\b",
        r"\b\d+(?:\.\d+)?\s*points?\b",
        r"\bpoints?\s*:?\s*\d+(?:\.\d+)?\b"
    ]

    return any(
        re.search(
            pattern,
            text,
            re.IGNORECASE
        )
        for pattern in patterns
    )



def inspect_assignment_page(
    driver,
    assignment_url,
    assumed_year
):
    """
    Visit an individual assignment page when the listing doesn't
    provide enough information.

    Returns:
        (due_date, graded)
    """

    original_window = driver.current_window_handle

    due_date = None
    graded = False

    try:

        driver.switch_to.new_window("tab")
        driver.get(assignment_url)

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located(
                (By.TAG_NAME, "body")
            )
        )


        time_elements = driver.find_elements(
            By.CSS_SELECTOR,
            "time"
        )

        for element in time_elements:

            try:

                datetime_value = element.get_attribute(
                    "datetime"
                )

                text_value = element.text

                if datetime_value:

                    try:

                        parsed = datetime.fromisoformat(
                            datetime_value.replace(
                                "Z",
                                "+00:00"
                            )
                        )

                        due_date = parsed.date()
                        break

                    except ValueError:
                        pass

                if text_value and due_date is None:

                    possible = extract_date_from_text(
                        text_value,
                        assumed_year
                    )

                    if possible:
                        due_date = possible
                        break

            except Exception:
                pass



        page_text = driver.find_element(
            By.TAG_NAME,
            "body"
        ).text

        if due_date is None:

            # Prioritize text after labels like Due.
            due_patterns = [
                rf"Due[^\n]*{MONTH_PATTERN}\s+\d{{1,2}}",
                rf"Available until[^\n]*{MONTH_PATTERN}\s+\d{{1,2}}"
            ]

            for pattern in due_patterns:

                match = re.search(
                    pattern,
                    page_text,
                    re.IGNORECASE
                )

                if match:

                    due_date = extract_date_from_text(
                        match.group(0),
                        assumed_year
                    )

                    if due_date:
                        break

        graded = has_points(page_text)

    except Exception:
        pass

    finally:

        try:
            driver.close()
        except Exception:
            pass

        try:
            driver.switch_to.window(
                original_window
            )
        except Exception:
            pass

    return due_date, graded


# SELENIUM SCRAPER


def selenium_scrape_missing(
    driver,
    selected_courses,
    existing_assignments,
    start_date,
    end_date
):
    """
    Perform the second Canvas pass.

    ICS data is considered the first source.

    Selenium adds only assignments that appear to be graded and
    were missing from the ICS feed.
    """

    print(
        "\nStarting Selenium second-pass check..."
    )

    existing_keys = {
        assignment_key(item)
        for item in existing_assignments
    }

    extra_assignments = []

    assumed_year = start_date.year

    for course_name, course_url in selected_courses:

        print(
            f"\nChecking {course_name}..."
        )

        course_candidates = {}


        for tab in COURSE_TABS:

            page_url = (
                course_url.rstrip("/")
                + "/"
                + tab
            )

            try:

                driver.get(page_url)

                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located(
                        (By.TAG_NAME, "body")
                    )
                )

                time.sleep(1)

                candidates = (
                    candidate_assignment_links(
                        driver
                    )
                )

                for candidate in candidates:

                    course_candidates[
                        candidate["href"]
                    ] = candidate

            except Exception as error:

                print(
                    f"  Could not fully read "
                    f"{tab}: {error}"
                )

        print(
            f"  Found {len(course_candidates)} "
            f"candidate items."
        )


        for candidate in course_candidates.values():

            name = candidate["name"]
            href = candidate["href"]
            context = candidate["context"]

            lower_name = name.lower()

            if any(
                skip_word in lower_name
                for skip_word in SKIP_WORDS
            ):
                continue

            due_date = extract_date_from_text(
                context,
                assumed_year
            )

            graded = has_points(
                context
            )

            if due_date is None:

                inner_due_date, inner_graded = (
                    inspect_assignment_page(
                        driver,
                        href,
                        assumed_year
                    )
                )

                if inner_due_date:
                    due_date = inner_due_date

                if inner_graded:
                    graded = True

            # Ignore non-graded content such as readings,
            # PowerPoints, etc.
            if not graded:
                continue

            # No usable due date.
            if due_date is None:
                continue

            # Outside requested date range.
            if not (
                start_date
                <= due_date
                <= end_date
            ):
                continue

            assignment = {
                "Assignment": name,
                "Course": course_name,
                "Due Date": due_date,
                "Notes": ""
            }

            key = assignment_key(
                assignment
            )

            # Already present in ICS.
            if key in existing_keys:
                continue

            extra_assignments.append(
                assignment
            )

            existing_keys.add(key)

    print(
        "\nSelenium found "
        f"{len(extra_assignments)} additional "
        "assignments that were not already in the ICS data."
    )

    return extra_assignments


# MERGE AND DUPLICATE REMOVAL


def merge_assignments(
    ics_assignments,
    selenium_assignments
):
    """
    Combine the two sources while removing duplicates.
    """

    merged = []
    seen = set()

    for item in (
        ics_assignments
        + selenium_assignments
    ):

        key = assignment_key(item)

        if key in seen:
            continue

        seen.add(key)
        merged.append(item)

    # Oldest date first.
    merged.sort(
        key=lambda item: item["Due Date"]
    )

    return merged


# PASTEL COURSE COLORS


def generate_pastel_colors(count):
    """
    Generate as many distinct pastel colors as necessary.

    This prevents course colors from repeating simply because
    a fixed palette ran out.
    """

    colors = []

    if count <= 0:
        return colors

    for index in range(count):

        hue = index / count

        saturation = 0.30
        value = 1.00

        red, green, blue = colorsys.hsv_to_rgb(
            hue,
            saturation,
            value
        )

        # Mix slightly with white.
        red = int((red * 0.65 + 0.35) * 255)
        green = int((green * 0.65 + 0.35) * 255)
        blue = int((blue * 0.65 + 0.35) * 255)

        colors.append(
            f"{red:02X}{green:02X}{blue:02X}"
        )

    return colors



# EXCEL EXPORT


def export_excel(
    assignments,
    filename
):
    """
    Export the final assignment list.

    Features:
      • Oldest due date first
      • Blank row between different dates
      • Course-specific pastel colors
      • Unknown courses remain uncolored
      • Wide Notes column
      • Auto-sized other columns
    """

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Assignments"

    headers = [
        "Assignment",
        "Course",
        "Due Date",
        "Notes"
    ]

    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = sheet.cell(
            row=1,
            column=column_number,
            value=header
        )

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            vertical="center"
        )

    # Create one unique pastel color per known course.

    known_courses = sorted({
        item["Course"]
        for item in assignments
        if normalize_text(item["Course"])
        not in {
            "unknown",
            "unknown course",
            ""
        }
    })

    pastel_colors = generate_pastel_colors(
        len(known_courses)
    )

    course_colors = dict(
        zip(
            known_courses,
            pastel_colors
        )
    )

    row_number = 2
    previous_date = None

    for assignment in assignments:

        due_date = assignment[
            "Due Date"
        ]

        # Blank row when the date changes.
        if (
            previous_date is not None
            and due_date != previous_date
        ):
            row_number += 1

        assignment_cell = sheet.cell(
            row=row_number,
            column=1,
            value=assignment["Assignment"]
        )

        course_cell = sheet.cell(
            row=row_number,
            column=2,
            value=assignment["Course"]
        )

        date_cell = sheet.cell(
            row=row_number,
            column=3,
            value=due_date
        )

        notes_cell = sheet.cell(
            row=row_number,
            column=4,
            value=""
        )

        # Actual Excel date instead of text.
        date_cell.number_format = (
            "mmmm dd, yyyy"
        )

        course_name = assignment[
            "Course"
        ]

        if course_name in course_colors:

            color = course_colors[
                course_name
            ]

            course_cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"
            )

        previous_date = due_date
        row_number += 1


    # Autosize Assignment / Course / Due Date.


    for column_number in range(1, 4):

        max_length = 0

        for row in sheet.iter_rows(
            min_col=column_number,
            max_col=column_number
        ):

            cell = row[0]

            if cell.value is None:
                continue

            if isinstance(
                cell.value,
                (date, datetime)
            ):
                displayed_length = 18
            else:
                displayed_length = len(
                    str(cell.value)
                )

            max_length = max(
                max_length,
                displayed_length
            )

        sheet.column_dimensions[
            get_column_letter(
                column_number
            )
        ].width = max_length + 2

    # Approximately 500 pixels.
    sheet.column_dimensions["D"].width = 70

    # Freeze headers.
    sheet.freeze_panes = "A2"

    workbook.save(
        filename
    )


# MAIN PROGRAM


def run_full_scrape():

    print(
        "\n"
        "========================================\n"
        " Canvas Assignment Organizer - Full Mode\n"
        "========================================"
    )


    # 1. ICS location.


    ics_source = input(
        "\nPaste your Canvas ICS calendar link "
        "or local .ics file path:\n> "
    ).strip()

 # 2. Date range.

    print(
        "\nEnter the assignment date range."
    )

    start_date = parse_user_date(
        "Start date (example: April 07 2025): "
    )

    end_date = parse_user_date(
        "End date   (example: April 13 2025): "
    )

    while end_date < start_date:

        print(
            "\nThe ending date cannot be earlier "
            "than the starting date."
        )

        end_date = parse_user_date(
            "End date: "
        )

    # 3. ICS first pass.

    ics_assignments = parse_ics_calendar(
        ics_source,
        start_date,
        end_date
    )

    # 4. Selenium.

    print(
        "\nStarting Canvas browser check..."
    )

    driver = create_driver()

    try:

        driver.get(
            CANVAS_BASE_URL
        )

        if not HEADLESS:

            print(
                "\nIf Canvas asks you to sign in, "
                "please do so in Chrome."
            )

            input(
                "Once you are on your Canvas Dashboard, "
                "press Enter here to continue..."
            )

        else:

            print(
                "\nHeadless mode is enabled."
            )

            print(
                "The Chrome profile configured at the top "
                "of the script must already be signed "
                "into Canvas."
            )

        # 5. Course selection.

        courses = get_canvas_courses(
            driver
        )

        selected_courses = select_courses(
            courses
        )

        if not selected_courses:

            print(
                "\nNo courses were selected."
            )

            return

        # 6. Selenium second pass.

        selenium_assignments = (
            selenium_scrape_missing(
                driver,
                selected_courses,
                ics_assignments,
                start_date,
                end_date
            )
        )

        # 7. Merge.

        assignments = merge_assignments(
            ics_assignments,
            selenium_assignments
        )

        if not assignments:

            print(
                "\nNo assignments were found "
                "for the requested date range."
            )

            return

        print(
            "\nFinal results:"
        )

        print(
            f"  ICS assignments: "
            f"{len(ics_assignments)}"
        )

        print(
            f"  Selenium additions: "
            f"{len(selenium_assignments)}"
        )

        print(
            f"  Final unique assignments: "
            f"{len(assignments)}"
        )

        # 8. Filename LAST.


        filename = clean_filename(
            input(
                "\nEnter the filename for the Excel "
                "spreadsheet:\n> "
            )
        )

        # 9. Excel.

        export_excel(
            assignments,
            filename
        )

        print(
            f"\nFile saved as:\n{filename}"
        )

        print(
            "\nThank you for using this python program, "
            "good luck with your classes!"
        )

    finally:

        try:
            driver.quit()
        except Exception:
            pass

# START PROGRAM


if __name__ == "__main__":
    run_full_scrape()
